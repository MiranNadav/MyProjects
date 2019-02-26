# -*- coding: utf-8 -*-

import os
import re
import codecs
import sys

# textfile = '/Users/nmiran/Desktop/nadav.txt'
textfile = '/Users/nmiran/Desktop/BenGurion.txt'
folderpath = '/Users/nmiran/Desktop/textfiles'

excelFilePath = '/Users/nmiran/Desktop/nadav.xlsx'
ignoreWordsEnglish = ['a', 'about',
                      'above',
                      'after',
                      'again',
                      'against',
                      'all',
                      'am',
                      'an',
                      'and',
                      'any',
                      'are',
                      'aren\'t',
                      'as',
                      'at',
                      'be',
                      'because',
                      'been',
                      'before',
                      'being',
                      'below',
                      'between',
                      'both',
                      'but',
                      'by',
                      'can\'t',
                      'cannot',
                      'could',
                      'couldn\'t',
                      'did',
                      'didn\'t',
                      'do',
                      'does',
                      'doesn\'t',
                      'doing',
                      'don\'t',
                      'down',
                      'during',
                      'each',
                      'few',
                      'for',
                      'from',
                      'further',
                      'had',
                      'hadn\'t',
                      'has',
                      'hasn\'t',
                      'have',
                      'haven\'t',
                      'having',
                      'he',
                      'he\'d',
                      'he\'ll',
                      'he\'s',
                      'her',
                      'here',
                      'here\'s',
                      'hers',
                      'herself',
                      'him',
                      'himself',
                      'his',
                      'how',
                      'how\'s',
                      'i',
                      'i\'d',
                      'i\'ll',
                      'i\'m',
                      'i\'ve',
                      'if',
                      'in',
                      'into',
                      'is',
                      'isn\'t',
                      'it',
                      'it\'s',
                      'its',
                      'itself',
                      'let\'s',
                      'me',
                      'more',
                      'most',
                      'mustn\'t',
                      'my',
                      'myself',
                      'no',
                      'nor',
                      'not',
                      'of',
                      'off',
                      'on',
                      'once',
                      'only',
                      'or',
                      'other',
                      'ought',
                      'our',
                      'ours',
                      'ourselves',
                      'out',
                      'over',
                      'own',
                      'same',
                      'shan\'t',
                      'she',
                      'she\'d',
                      'she\'ll',
                      'she\'s',
                      'should',
                      'shouldn\'t',
                      'so',
                      'some',
                      'such',
                      'than',
                      'that',
                      'that\'s',
                      'the',
                      'their',
                      'theirs',
                      'them',
                      'themselves',
                      'then',
                      'there',
                      'there\'s',
                      'these',
                      'they',
                      'they\'d',
                      'they\'ll',
                      'they\'re',
                      'they\'ve',
                      'this',
                      'those',
                      'through',
                      'to',
                      'too',
                      'under',
                      'until',
                      'up',
                      'very',
                      'was',
                      'wasn\'t',
                      'we',
                      'we\'d',
                      'we\'ll',
                      'we\'re',
                      'we\'ve',
                      'were',
                      'weren\'t',
                      'what',
                      'what\'s',
                      'when',
                      'when\'s',
                      'where',
                      'where\'s',
                      'which',
                      'while',
                      'who',
                      'who\'s',
                      'whom',
                      'why',
                      'why\'s',
                      'with',
                      'won\'t',
                      'would',
                      'wouldn\'t',
                      'you',
                      'you\'d',
                      'you\'ll',
                      'you\'re',
                      'you\'ve', 'your',
                      'yours',
                      'yourself',
                      'yourselves', ]
ignoreWordsHebrew = [r'אבל',
                     r'או',
                     r'אולי',
                     r'אותה',
                     r'אותו',
                     r'אותי',
                     r'אותך',
                     r'אותם',
                     r'אותן',
                     r'אותנו',
                     r'אז',
                     r'אחר',
                     r'אחרות',
                     r'אחרי',
                     r'אחריכן',
                     r'אחרים',
                     r'אחרת',
                     r'אי',
                     r'איזה',
                     r'איך',
                     r'אין',
                     r'איפה',
                     r'איתה',
                     r'איתו',
                     r'איתי',
                     r'איתך',
                     r'איתכם',
                     r'איתכן',
                     r'איתם',
                     r'איתן',
                     r'איתנו',
                     r'אך',
                     r'אל',
                     r'אלה',
                     r'אלו',
                     r'אם',
                     r'אנחנו',
                     r'אני',
                     r'אס',
                     r'אף',
                     r'אצל',
                     r'אשר',
                     r'את',
                     r'אתה',
                     r'אתכם',
                     r'אתכן',
                     r'אתם',
                     r'אתן',
                     r'באיזומידה',
                     r'באמצע',
                     r'באמצעות',
                     r'בגלל',
                     r'בין',
                     r'בלי',
                     r'במידה',
                     r'במקוםשבו',
                     r'ברם',
                     r'בשביל',
                     r'בשעהש',
                     r'בתוך',
                     r'גם',
                     r'דרך',
                     r'הוא',
                     r'היא',
                     r'היה',
                     r'היכן',
                     r'היתה',
                     r'היתי',
                     r'הם',
                     r'הן',
                     r'הנה',
                     r'הסיבהשבגללה',
                     r'הרי',
                     r'ואילו',
                     r'ואת',
                     r'זאת',
                     r'זה',
                     r'זות',
                     r'יהיה',
                     r'יוכל',
                     r'יוכלו',
                     r'יותרמדי',
                     r'יכול',
                     r'יכולה',
                     r'יכולות',
                     r'יכולים',
                     r'יכל',
                     r'יכלה',
                     r'יכלו',
                     r'יש',
                     r'כאן',
                     r'כאשר',
                     r'כולם',
                     r'כולן',
                     r'כזה',
                     r'כי',
                     r'כיצד',
                     r'כך',
                     r'ככה',
                     r'כל',
                     r'כלל',
                     r'כמו',
                     r'כן',
                     r'כפי',
                     r'כש',
                     r'לא',
                     r'לאו',
                     r'לאיזותכלית',
                     r'לאן',
                     r'לבין',
                     r'לה',
                     r'להיות',
                     r'להם',
                     r'להן',
                     r'לו',
                     r'לי',
                     r'לכם',
                     r'לכן',
                     r'למה',
                     r'למטה',
                     r'למעלה',
                     r'למקוםשבו',
                     r'למרות',
                     r'לנו',
                     r'לעבר',
                     r'לעיכן',
                     r'לפיכך',
                     r'לפני',
                     r'מאד',
                     r'מאחורי',
                     r'מאיזוסיבה',
                     r'מאין',
                     r'מאיפה',
                     r'מבלי',
                     r'מבעד',
                     r'מדוע',
                     r'מה',
                     r'מהיכן',
                     r'מול',
                     r'מחוץ',
                     r'מי',
                     r'מכאן',
                     r'מכיוון',
                     r'מלבד',
                     r'מן',
                     r'מנין',
                     r'מסוגל',
                     r'מעט',
                     r'מעטים',
                     r'מעל',
                     r'מצד',
                     r'מקוםבו',
                     r'מתחת',
                     r'מתי',
                     r'נגד',
                     r'נגר',
                     r'נו',
                     r'עד',
                     r'עז',
                     r'על',
                     r'עלי',
                     r'עליה',
                     r'עליהם',
                     r'עליהן',
                     r'עליו',
                     r'עליך',
                     r'עליכם',
                     r'עלינו',
                     r'עם',
                     r'עצמה',
                     r'עצמהם',
                     r'עצמהן',
                     r'עצמו',
                     r'עצמי',
                     r'עצמם',
                     r'עצמן',
                     r'עצמנו',
                     r'פה',
                     r'רק',
                     r'שוב',
                     r'של',
                     r'שלה',
                     r'שלהם',
                     r'שלהן',
                     r'שלו',
                     r'שלי',
                     r'שלך',
                     r'שלכה',
                     r'שלכם',
                     r'שלכן',
                     r'שלנו',
                     r'שם',
                     r'תהיה',
                     r'תחת']
ignoreSigns = ['\'', '-', '(', ')', ':', ';', '.', ',', '\\', '/', '?', '!']


def replaceSigns(data, words, sign):
    for word in words:
        data = data.replace(word, sign)
    return data


def replaceWords(data, words, sign):
    for word in words:
        data = re.sub(r'\b%s\b' % word, sign, data)
        # data = data.replace(word, sign)
    return data


def replaceWordsHebrew(data, words, sign):
    # data = data.encode('utf-8')
    for word in words:
        data = re.sub(r' %s ' % word, sign, data)
    return data


def countOccurrences(word, text):
    # checks if "word" contains a space (meaning it's a phrase)
    if ' ' in word:
        return text.lower().count(word.lower())
    else:
        return text.lower().split().count(word.lower())


def listOfOccurences(wordList, text):
    list = []
    for word in wordList:
        list.append((word, countOccurrences(word, text)))
    return list

def listOfOccurences2(wordsList, text):
  list = []
  for word in wordsList:
    list.append((word, countOccurences2(word,text)))
  return list

def countOccurences2 (partialWord, text):
  wordsList = text.lower().split()
  count = 0
  for word in wordsList:
    if partialWord in word:
      count = count + 1 
  return count

def getValidInputWord():
    try:
        input_from_user = input(
            'Please enter words you want to find - either one word or a list of words seperated by space\n')
        return input_from_user
    except:
        print('Wrong Input, please enter a valid String!')
        return getValidInputWord()


def getValidPhrase():
    try:
        input_from_user = input(
            'Please enter a phrase you want to find\n')
        return input_from_user
    except:
        print('Wrong Input, please enter a valid String!')
        return getValidInputWord()


def getValidFilePath():
    try:
        input_from_user = input('Please enter a path for the txt file\n')
        return input_from_user
    except:
        print('Illegal input, please enter a valid path!')
        return getValidFilePath()


def getValidFolderPath():
    try:
        input_from_user = input('Please enter a folder path:\n')
        # if input_from_user == "":
        #     return
        if os.path.isdir(input_from_user):
            return input_from_user
        else:
            print('This is not a folder, please enter a valid path or press Enter to exit')
            return getValidFolderPath()
    except:
        print('Illegal input, please enter a valid path!')
        return getValidFolderPath()


def receiveWordsToFind():
    listOfWords = []
    print ('dont forget to enter **\'stop\'** to finish')
    inputWord = getValidInputWord()

    while inputWord != -1 and inputWord.lower() != 'stop':
        try:
            if inputWord.index(' ') > 0:
                currentWords = inputWord.replace(" ", " ").split(" ")
                for word in currentWords:
                    listOfWords.append(word)
                inputWord = getValidInputWord()
        except:
            listOfWords.append(inputWord)
            inputWord = getValidInputWord()

    listOfWords = receivePhrasesToFind(listOfWords)
    return listOfWords


def receivePhrasesToFind(listOfWords):
    print (
        'Now, enter full phrases (full string - for example "Hello World"). dont forget to enter **\'stop\'** to end the script')
    inputPhrase = getValidPhrase()
    while inputPhrase != 'stop':
        listOfWords.append(inputPhrase)
        inputPhrase = getValidPhrase()
    return listOfWords


def removeDups(duplicate):
    final_list = []
    for word in duplicate:
        word = word.lower()
        if word not in final_list:
            final_list.append(word)
    return final_list


def writeToExcelFile(findingsList, file_name, name):
    import openpyxl
    import os
    import datetime

    now = datetime.datetime.now()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws['A1'] = textfile
    ws['A2'] = 'Word'
    ws['B2'] = 'Num Of Occurences'

    # print(findingsList)

    for word in findingsList:
        ws.append(word)

    excel_path_name = "/" + file_name + '-Word-occurences-' + name + "-" + now.strftime('%d-%m-%y-%H%M') + '.xlsx'
    path_to_save = os.getcwd() + excel_path_name
    wb.save(path_to_save)


def openFile():
    text_file_path = getValidFilePath()
    try:
        # with open(text_file_path, 'r') as myfile:
        with codecs.open(text_file_path, "r", encoding='utf-8', errors='ignore') as myfile:
            all_words = myfile.read().replace('\n', ' ')
        return all_words
    except:
        return openFile()


def openFileFromFolder(file):
    return file.read().replace('\n', ' ')


def getWordsList(text):
    # return re.compile('\w+').findall(text)
    return re.split('\s+', text)


def removeErrorWords(word_list):
    clean_list = []
    for word in word_list:
        if len(word) > 1 and not word.isdigit():
            clean_list.append(word)
    return clean_list


def mainProcessor(all_words, file_name, words_to_find):
    # Clear unnecessary signs and words (StopWords in Hebrew and English)
    all_words = all_words.lower()
    all_words = replaceSigns(all_words,
                             ignoreSigns, ' ')
    all_words = replaceWords(all_words, ignoreWordsEnglish, ' ')
    all_words = replaceWordsHebrew(all_words, ignoreWordsHebrew, ' ')
    # all_words_list = re.findall(r'[^\s!,.?":;0-9]+', all_words)
    all_words = " ".join(all_words.split())

    # Create a clean word lists (all words in file, excluding StopWords)
    all_words_list = getWordsList(all_words)
    [word.lower() for word in all_words_list]
    all_words_list = removeDups(all_words_list)
    all_words_list = removeErrorWords(all_words_list)

    # Create Finding lists from the given words
    findingsList = listOfOccurences(words_to_find, all_words)
    findingsList.sort(key=lambda tup: tup[1])
    findingsList = list(reversed(findingsList))

    # Create Finding list from all words in file (excluding StopWords)
    findingslist_unfiltered = listOfOccurences(all_words_list, all_words)
    findingslist_unfiltered.sort(key=lambda tup: tup[1])
    findingslist_unfiltered = list(reversed(findingslist_unfiltered))

    # Create Findings list from regex of given words
    findingsListContains = listOfOccurences2(words_to_find, all_words)
    findingsListContains.sort(key=lambda tup: tup[1])
    findingsListContains = list(reversed(findingsListContains))



    # Write 2 excel files - filtered & unfiltered
    writeToExcelFile(findingslist_unfiltered, file_name, 'All')
    writeToExcelFile(findingsList, file_name, 'Filtered')
    writeToExcelFile(findingsListContains, file_name,'Contains')


def main():
    user_selection = 0
    while user_selection != 1 and user_selection != 2:
        try:
            user_selection = int(input("Press 1 to enter file path, and 2 for a folder:\n"))
        except:
            continue

    # Work on a file
    if user_selection == 1:
        file = openFile()
        words_to_find = removeDups(receiveWordsToFind())
        mainProcessor(file, "", words_to_find)
    # Work on a directory
    else:
        folder_path = getValidFolderPath()
        words_to_find = removeDups(receiveWordsToFind())

        files = []
        for file_path in os.listdir(folder_path):
            if file_path.endswith('.txt'):
                file = codecs.open(folder_path + "/" + file_path, "r", encoding='utf-8', errors='ignore')
                files.append(file)

        for file in files:
            file_base_name = os.path.basename(file.name)
            file_base_name = os.path.splitext(file_base_name)[0]
            all_words = openFileFromFolder(file)
            mainProcessor(all_words, file_base_name, words_to_find)


main()
